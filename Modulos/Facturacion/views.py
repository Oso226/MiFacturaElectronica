# ======================================================
# 📦 IMPORTACIONES GLOBALES
# ======================================================
import os
import io
import qrcode
import json
import time
import base64
import uuid
import tempfile
import threading
import openpyxl
from io import BytesIO  # ✅ agregado correctamente

from decimal import Decimal
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import IntegrityError
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.utils.timezone import now, localtime
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.template.loader import render_to_string
from django.conf import settings
from django.core.mail import EmailMessage, send_mail
from openpyxl.styles import Alignment, Font, Border, Side
from weasyprint import HTML, CSS
from django.db.models import Max

# Modelos y formularios
from .models import (
    DTE, Perfil, Cliente, Proveedor, Producto,
    Inventario, Compra, DetalleDTE, Empresa
)
from .forms import ClienteForm, ProveedorForm, ProductoForm
from .permisos import rol_requerido
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required


# ======================================================
# 🔐 INICIAR SESIÓN (versión definitiva y segura)
# ======================================================
def iniciar_sesion(request):
    """
    Inicia sesión de usuario verificando rol y empresa asociada.
    Muestra mensaje post-logout de forma segura.
    """

    # 🟢 Mostrar mensaje si viene desde logout
    if request.GET.get("logout") == "1":
        # Evita error si storage aún no existe
        try:
            messages.success(request, "Sesión cerrada correctamente.")
        except Exception:
            pass

    # 🧹 Limpia mensajes antiguos solo si el middleware está activo
    if hasattr(request, '_messages'):
        storage = messages.get_messages(request)
        for _ in storage:
            pass  # vacía sin borrar el backend

    # 🔐 Si ya está logueado, ir al menú
    if request.user.is_authenticated:
        return redirect('menu_principal')

    # 🧩 Si envía el formulario de login
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            perfil = Perfil.objects.filter(user=user).select_related('empresa').first()

            if not perfil:
                logout(request)
                messages.error(request, "El usuario no tiene un perfil asignado.")
                return redirect('login')

            if not perfil.empresa:
                messages.warning(request, "No se encontró empresa asignada al usuario.")
            else:
                request.session['empresa_id'] = perfil.empresa.id

            messages.success(request, f"Inicio de sesión exitoso como {perfil.rol}.")
            return redirect('menu_principal')
        else:
            messages.error(request, "Usuario o contraseña incorrectos.")

    return render(request, 'registration/login.html')


# ======================================================
# 🚪 CERRAR SESIÓN
# ======================================================
def cerrar_sesion(request):
    """
    Cierra la sesión y redirige limpiamente al login
    evitando errores en el sistema de mensajes.
    """
    logout(request)
    # En lugar de usar messages aquí, mandamos un parámetro en la URL
    return redirect('/login/?logout=1')


def registro(request):
    """
    Registra un nuevo usuario con su rol y crea automáticamente el perfil.
    Por seguridad, puede deshabilitarse el acceso público en producción.
    """
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password1')
        rol = request.POST.get('rol')
        empresa_id = request.POST.get('empresa')  # opcional si se asigna desde admin

        if User.objects.filter(username=username).exists():
            messages.error(request, "El nombre de usuario ya existe.")
        else:
            user = User.objects.create_user(username=username, password=password)
            perfil = Perfil.objects.create(user=user, rol=rol)

            # 🔹 Si se envía empresa desde el formulario
            if empresa_id:
                from .models import Empresa
                empresa = Empresa.objects.filter(id=empresa_id).first()
                if empresa:
                    perfil.empresa = empresa
                    perfil.save()

            messages.success(request, f"Usuario '{username}' registrado exitosamente como {rol}.")
            return redirect('login')

    return render(request, 'registration/registro.html')

# ======================================================
# 🧭 MENÚ PRINCIPAL
# ======================================================

@login_required
def menu_principal(request):
    """
    Muestra el menú principal con el rol y empresa del usuario logueado.
    No genera error si el perfil no tiene empresa asignada.
    """
    perfil = Perfil.objects.select_related('empresa').filter(user=request.user).first()

    # Si no hay perfil asociado al usuario
    if not perfil:
        messages.error(request, "No se encontró un perfil asociado al usuario.")
        return redirect('logout')

    # Si el perfil no tiene empresa, evitar error 500
    if not perfil.empresa:
        empresa_nombre = "Sin empresa asignada"
    else:
        empresa_nombre = perfil.empresa.nombre
        request.session['empresa_id'] = perfil.empresa.id  # Guardar en sesión

    # 🔹 Contexto enviado al template
    contexto = {
        'rol': perfil.rol,
        'empresa': empresa_nombre
    }

    print("Rol:", perfil.rol)
    print("Empresa mostrada:", empresa_nombre)

    return render(request, 'Facturacion/menu_principal.html', contexto)

# 📄 Lista de documentos emitidos
@rol_requerido(['Administrador', 'Contador', 'Empleado'])
def lista_dte(request):
    dtes = DTE.objects.all().order_by('-fecha_emision')
    return render(request, 'Facturacion/lista_dte.html', {'dtes': dtes})


# 🧾 Crear nuevo documento
@rol_requerido(['Administrador', 'Empleado'])
def crear_dte(request, tipo_dte=None):
    if not tipo_dte:
        messages.warning(request, "Debes seleccionar un tipo de documento.")
        return redirect('menu_facturacion')

    # 🔹 Ajuste de rutas de plantillas
    plantillas_dte = {
        '01': 'Facturacion/factura01.html',
        '03': 'Facturacion/ccf03.html',
        '05': 'Facturacion/notaCredito05.html',
        '06': 'Facturacion/notaDebito06.html',
        '07': 'Facturacion/retencion07.html',
        '11': 'Facturacion/liquidacion11.html',
    }

    template = plantillas_dte.get(tipo_dte, 'Facturacion/factura01.html')
    return render(request, template, {'tipo_dte': tipo_dte})


# ❌ Anular documento
@rol_requerido(['Administrador', 'Contador'])
def anular_dte(request, id):
    dte = get_object_or_404(DTE, id=id)
    dte.estado = 'Anulado'
    dte.save()
    messages.success(request, f'Documento {dte.numero_control} ha sido anulado correctamente.')
    return redirect('lista_dte')


# 📊 Reporte de ventas
@rol_requerido(['Administrador', 'Contador'])
def reporte_ventas(request):
    """
    Muestra los últimos 10 comprobantes emitidos, 
    ordenados del más reciente al más antiguo.
    """
    dtes = DTE.objects.filter(estado='Activo').order_by('-fecha_emision')[:10]
    total_general = sum(d.total for d in dtes)

    return render(request, 'Facturacion/ventas.html', {
        'dtes': dtes,
        'total': total_general
    })


# 🧭 Menú principal de facturación
@login_required
def menu_facturacion(request):
    """ Muestra menú general de facturación con diseño coherente """
    perfil = getattr(request.user, 'perfil', None)
    rol = perfil.rol if perfil else 'Sin Rol'

    return render(request, 'Facturacion/menu_facturacion.html', {
        'rol': rol,
        'empresa': getattr(perfil, 'empresa', 'Sin empresa asignada')
    })


# 🧾 Vista individual de factura tipo 01
@login_required
def nueva_factura(request, tipo):
    dte = get_object_or_404(DTE, tipo_dte=tipo)
    return render(request, 'Facturacion/factura01.html', {
        'dte': dte,
        'empresa': dte.empresa,
        'cliente': dte.cliente
    })


# ======================================================
# ⚙️ GENERAR DTE (Simulación de envío con PDF y JSON usando SendGrid API)
# ======================================================
@csrf_exempt
@login_required
def generar_dte(request, tipo):
    """
    Simula el envío del DTE al Ministerio y envía al cliente:
    - El aviso de aprobación.
    - Adjuntos: comprobante en PDF y JSON.
    """
    import sendgrid
    from sendgrid.helpers.mail import Mail, Email, To, Content, Attachment, FileContent, FileName, FileType, Disposition
    import base64

    try:
        dte = DTE.objects.filter(tipo_dte=tipo).last()
        if not dte:
            return JsonResponse({'success': False, 'error': 'No se encontró el documento.'})

        # ⏳ Simulación del proceso de generación
        time.sleep(2)
        dte.codigo_generacion = f"MH-{int(time.time())}"
        dte.sello_recepcion = f"SELLO-{int(time.time())}"
        dte.save()

        # =====================================================
        # 🔹 Crear JSON temporal (compatible con Render)
        # =====================================================
        temp_dir = tempfile.gettempdir()
        json_filename = f"DTE_{dte.numero_control}.json"
        json_path = os.path.join(temp_dir, json_filename)

        dte_json = {
            "tipo_dte": dte.tipo_dte,
            "numero_control": dte.numero_control,
            "cliente": dte.cliente.nombre if dte.cliente else "Sin cliente",
            "correo": dte.cliente.correo if dte.cliente else "",
            "fecha_emision": dte.fecha_emision.strftime("%Y-%m-%d"),
            "total": float(dte.total),
            "codigo_generacion": dte.codigo_generacion,
            "sello_recepcion": dte.sello_recepcion,
        }

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(dte_json, f, ensure_ascii=False, indent=4)

        # =====================================================
        # 🔹 Generar PDF con WeasyPrint
        # =====================================================
        html_content = render_to_string("Facturacion/factura01.html", {
            "dte": dte,
            "empresa": dte.empresa,
            "cliente": dte.cliente,
            "editable": False,
            "qr_data": None,
        })

        # ✅ Opción 2: Ruta automática para desarrollo y producción
        dev_css = os.path.join(settings.BASE_DIR, 'Modulos', 'Facturacion', 'static', 'css', 'factura01.css')
        prod_css = os.path.join(settings.BASE_DIR, 'staticfiles_dev', 'css', 'factura01.css')

        if os.path.exists(prod_css):
            css_path = prod_css
            print(f"📁 Usando CSS de producción: {css_path}")
        else:
            css_path = dev_css
            print(f"📁 Usando CSS local: {css_path}")

        pdf_buffer = BytesIO()
        HTML(string=html_content).write_pdf(pdf_buffer, stylesheets=[CSS(filename=css_path)])
        pdf_buffer.seek(0)

        # =====================================================
        # 🔹 Envío de correo con SendGrid API
        # =====================================================
        if dte.cliente and dte.cliente.correo:
            print(f"🚀 Enviando correo a: {dte.cliente.correo}")

            sg = sendgrid.SendGridAPIClient(api_key=os.environ.get("SENDGRID_API_KEY"))

            # Configuración básica del correo
            from_email = Email("manuelito2327@gmail.com", "OMNIGEST Facturación Electrónica")
            to_email = To(dte.cliente.correo)
            subject = "📄 Comprobante Electrónico Aprobado - OMNIGEST"
            content = Content(
                "text/plain",
                f"""Estimado {dte.cliente.nombre},

Su documento electrónico tipo {dte.get_tipo_dte_display()} ha sido aprobado por el Ministerio de Hacienda.

Adjunto encontrará su comprobante en formato PDF y JSON.

Saludos cordiales,
Equipo OMNIGEST"""
            )

            mail = Mail(from_email, to_email, subject, content)

            # 📎 Adjuntar PDF
            pdf_data = base64.b64encode(pdf_buffer.read()).decode()
            mail.add_attachment(
                Attachment(
                    FileContent(pdf_data),
                    FileName(f"DTE_{dte.numero_control}.pdf"),
                    FileType("application/pdf"),
                    Disposition("attachment")
                )
            )

            # 📎 Adjuntar JSON
            with open(json_path, "rb") as jf:
                json_data = base64.b64encode(jf.read()).decode()
                mail.add_attachment(
                    Attachment(
                        FileContent(json_data),
                        FileName(json_filename),
                        FileType("application/json"),
                        Disposition("attachment")
                    )
                )

            # 📤 Enviar correo con SendGrid
            response = sg.client.mail.send.post(request_body=mail.get())

            if response.status_code in [200, 202]:
                print("✅ Correo enviado correctamente con SendGrid API.")
            else:
                print(f"❌ Error SendGrid: {response.status_code} -> {response.body}")

        return JsonResponse({'success': True, 'msg': '✅ DTE enviado exitosamente con adjuntos.'})

    except Exception as e:
        print(f"❌ Error general en generar_dte: {e}")
        return JsonResponse({'success': False, 'error': str(e)})

# ======================================================
# ✏️ EDITAR CAMPOS DEL CLIENTE EN UN DTE
# ======================================================
@rol_requerido(['Administrador', 'Gerente'])
@csrf_exempt
def editar_datos_dte(request, id):
    """
    Permite modificar datos del cliente (nombre, correo, NIT, dirección)
    y reenviar el comprobante actualizado (PDF + JSON) al cliente.
    """
    dte = get_object_or_404(DTE, id=id)

    if request.method == "POST":
        try:
            import json, tempfile
            data = json.loads(request.body.decode('utf-8'))

            # 🔹 Actualizar datos del cliente
            cliente = dte.cliente
            cliente.nombre = data.get('nombre', cliente.nombre)
            cliente.correo = data.get('correo', cliente.correo)
            cliente.nit = data.get('nit', cliente.nit)
            cliente.direccion = data.get('direccion', cliente.direccion)
            cliente.save()

            # 🔹 Simular nuevo código MH
            time.sleep(2)
            dte.codigo_generacion = "MH-" + str(int(time.time()))
            dte.sello_recepcion = "SELLO-" + str(int(time.time()))
            dte.save()

            # 🔹 Crear JSON actualizado
            json_path = tempfile.gettempdir() + f"\\DTE_{dte.numero_control}.json"

            dte_json = {
                "tipo_dte": dte.tipo_dte,
                "numero_control": dte.numero_control,
                "cliente": cliente.nombre,
                "correo": cliente.correo,
                "fecha_emision": dte.fecha_emision.strftime("%Y-%m-%d"),
                "total": float(dte.total),
                "codigo_generacion": dte.codigo_generacion,
                "sello_recepcion": dte.sello_recepcion,
            }

            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(dte_json, f, ensure_ascii=False, indent=4)

            # 🔹 Crear PDF actualizado con factura01.html
            from io import BytesIO
            from django.template.loader import render_to_string
            from xhtml2pdf import pisa

            pdf_buffer = BytesIO()
            # 📄 Renderizar la factura con tu plantilla HTML original
            html_content = render_to_string("Facturacion/factura01.html", {
                 "dte": dte,
                "empresa": dte.empresa,
                "cliente": dte.cliente,
                "editable": False,
                "qr_data": None,
            })

            # 🎨 Agregar tus estilos CSS reales
            css_path = os.path.join(settings.BASE_DIR, 'Modulos', 'Facturacion', 'static', 'css', 'factura01.css')
            pdf_buffer = BytesIO()

            # 🖨️ Generar PDF con WeasyPrint (idéntico a lo que ves en el navegador)
            HTML(string=html_content, base_url=request.build_absolute_uri("/")).write_pdf(
            pdf_buffer, stylesheets=[CSS(filename=css_path)]
            )

            pdf_buffer.seek(0)

            # 🔹 Enviar correo al cliente con los adjuntos
            if cliente.correo:
                from django.core.mail import EmailMessage

                email = EmailMessage(
                    subject="📄 DTE Actualizado y Aprobado - OMNIGEST",
                    body=(
                        f"Estimado {cliente.nombre},\n\n"
                        "Su documento electrónico ha sido actualizado y aprobado "
                        "por el Ministerio de Hacienda.\n\n"
                        "Adjunto encontrará su comprobante actualizado en PDF y JSON.\n\n"
                        "Saludos,\nEquipo OMNIGEST"
                    ),
                    from_email="facturacion@omnigest.com",
                    to=[cliente.correo],
                )

                # Adjuntar PDF actualizado
                email.attach(f"DTE_{dte.numero_control}.pdf", pdf_buffer.read(), "application/pdf")

                # Adjuntar JSON actualizado
                with open(json_path, "r", encoding="utf-8") as f:
                    email.attach(f"DTE_{dte.numero_control}.json", f.read(), "application/json")

                email.send(fail_silently=False)

            return JsonResponse({
                "status": "ok",
                "msg": "Documento actualizado y reenviado al cliente con comprobantes adjuntos."
            })

        except Exception as e:
            return JsonResponse({"status": "error", "msg": str(e)})

    return JsonResponse({"status": "error", "msg": "Método no permitido."})


@login_required
def ver_dte_json(request, id):
    """Devuelve los datos del DTE en formato JSON para edición rápida"""
    dte = get_object_or_404(DTE, id=id)
    return JsonResponse({
        "id": dte.id,
        "tipo_dte": dte.tipo_dte,
        "numero_control": dte.numero_control,
        "fecha": dte.fecha_emision.strftime("%Y-%m-%d %H:%M"),
        "estado": dte.estado,
        "total": float(dte.total),
        "cliente_nombre": dte.cliente.nombre if dte.cliente else "",
        "cliente_correo": dte.cliente.correo if dte.cliente else "",
        "cliente_nit": dte.cliente.nit if dte.cliente else "",
        "cliente_direccion": dte.cliente.direccion if dte.cliente else ""
    })

# ======================================================
# 📚 CATÁLOGOS
# ======================================================

@login_required
def menu_catalogos(request):
    return render(request, 'Facturacion/menu_catalogos.html')


@rol_requerido(['Administrador', 'Empleado'])
def lista_clientes(request):
    """
    Lista los clientes con opción de filtrar por nombre o DUI.
    Si no hay búsqueda, muestra los últimos 20 registros.
    """
    query = request.GET.get('q', '').strip()

    if query:
        # 🔹 Filtra por nombre o DUI
        clientes = Cliente.objects.filter(
            Q(nombre__icontains=query) | Q(dui__icontains=query)
        ).order_by('-id')
    else:
        # 🔹 Si no hay búsqueda, muestra los últimos 20
        clientes = Cliente.objects.all().order_by('-id')[:20]

    context = {
        'clientes': clientes,
        'query': query,  # Mantiene el texto del buscador
    }
    return render(request, 'Facturacion/clientes.html', context)


@rol_requerido(['Administrador', 'Empleado'])
def crear_cliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})
        return JsonResponse({'error': 'Formulario inválido'}, status=400)
    else:
        form = ClienteForm()
        return render(request, 'Facturacion/form_cliente.html', {'form': form})


@rol_requerido(['Administrador', 'Contador'])
def editar_cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id)

    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)  # ✅ Ojo aquí, instance=cliente
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': form.errors.as_json()})

    # Si es GET, muestra el formulario dentro del modal
    form = ClienteForm(instance=cliente)
    return render(request, 'Facturacion/editar_cliente.html', {'form': form, 'cliente': cliente})


@rol_requerido(['Administrador'])
def eliminar_cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id)
    cliente.delete()
    messages.success(request, "Cliente eliminado correctamente.")
    return redirect('lista_clientes')


# ======================================================
# 🏢 PROVEEDORES
# ======================================================

@login_required
def lista_proveedores(request):
    """
    Lista los proveedores con opción de filtrar por nombre o NIT.
    Si no hay búsqueda, muestra los últimos 20 registros.
    """
    query = request.GET.get('q', '').strip()

    if query:
        proveedores = Proveedor.objects.filter(
            Q(nombre__icontains=query) | Q(nit__icontains=query)
        ).order_by('-id')
    else:
        proveedores = Proveedor.objects.all().order_by('-id')[:20]

    return render(request, 'Facturacion/proveedores.html', {
        'proveedores': proveedores,
        'query': query,
    })


@login_required
def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})
        return JsonResponse({'error': 'Formulario inválido'}, status=400)
    else:
        form = ProveedorForm()
        return render(request, 'Facturacion/form_proveedor.html', {'form': form})


@login_required
def editar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)

    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': form.errors.as_json()})

    form = ProveedorForm(instance=proveedor)
    return render(request, 'Facturacion/editar_proveedor.html', {'form': form, 'proveedor': proveedor})

@login_required
def eliminar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    proveedor.delete()
    messages.success(request, f"Proveedor '{proveedor.nombre}' eliminado correctamente.")
    return redirect('lista_proveedores')

# ======================================================
# 📦 PRODUCTOS E INVENTARIO
# ======================================================

from django.db.models import Max, Q
from django.http import JsonResponse

@login_required
def lista_productos(request):
    query = request.GET.get('q', '')
    productos = Producto.objects.all().order_by('codigo')
    if query:
        productos = productos.filter(
            Q(codigo__icontains=query) |
            Q(descripcion__icontains=query)
        )
    return render(request, 'Facturacion/productos.html', {
        'productos': productos,
        'query': query
    })


def generar_codigo_producto():
    ultimo = Producto.objects.aggregate(ultimo=Max('codigo'))['ultimo']
    if ultimo:
        try:
            numero = int(ultimo.replace('PRD', '')) + 1
        except:
            numero = 1
    else:
        numero = 1
    return f"PRD{numero:04d}"


@login_required
def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save(commit=False)
            # Genera código único automáticamente
            producto.codigo = generar_codigo_producto()
            producto.inventario = 0
            producto.save()
            return JsonResponse({'success': True})
        return JsonResponse({'error': 'Formulario inválido'}, status=400)
    else:
        form = ProductoForm()
        # Autogenera el código en el formulario (solo visual)
        nuevo_codigo = generar_codigo_producto()
        return render(request, 'Facturacion/form_producto.html', {
            'form': form,
            'nuevo_codigo': nuevo_codigo
        })


@login_required
def editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)

    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            # ✅ Asegura que no modifique el código y que no falle por unique
            producto_editado = form.save(commit=False)
            producto_editado.codigo = producto.codigo  # mantiene el mismo código
            producto_editado.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'error': form.errors.as_json()}, status=400)

    # GET → cargar el formulario
    form = ProductoForm(instance=producto)
    return render(request, 'Facturacion/editar_producto.html', {
        'form': form,
        'producto': producto
    })

@login_required
def eliminar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    producto.delete()
    messages.success(request, f"Producto '{producto.descripcion}' eliminado correctamente.")
    return redirect('inventario')


# ======================================================
# 💰 COMPRAS E INVENTARIO
# ======================================================

@login_required
def inventario(request):
    query = request.GET.get('q', '')
    productos = Producto.objects.all().order_by('codigo')
    if query:
        productos = productos.filter(Q(codigo__icontains=query) | Q(descripcion__icontains(query)))
    return render(request, 'Facturacion/inventario.html', {'productos': productos, 'query': query})

def generar_comprobante_unico():
    """Genera un número único tipo COMP-20251113-0001"""
    ultimo = Compra.objects.aggregate(max_id=Max('id'))['max_id'] or 0
    return f"COMP-{now().strftime('%Y%m%d')}-{(ultimo + 1):04d}"


def generar_registro_unico():
    """Genera un número único tipo REG-00001"""
    ultimo = Compra.objects.aggregate(max_id=Max('id'))['max_id'] or 0
    return f"REG-{(ultimo + 1):05d}"


@login_required
def registrar_compra(request):
    if request.method == 'POST':
        try:
            # 🧩 Obtener datos del formulario
            producto_id = request.POST.get('producto')
            cantidad = int(request.POST.get('cantidad'))
            proveedor_id = request.POST.get('proveedor')
            precio_compra_str = request.POST.get('precio_compra')

            # ⚠️ Validaciones básicas
            if not producto_id or not proveedor_id:
                messages.error(request, "⚠️ Debes seleccionar un producto y un proveedor.")
                return redirect('registrar_compra')

            if not precio_compra_str or Decimal(precio_compra_str) <= 0:
                messages.error(request, "⚠️ Debes ingresar un precio total de compra válido.")
                return redirect('registrar_compra')

            # Convertir a Decimal
            precio_compra = Decimal(precio_compra_str)

            # Buscar producto y proveedor
            producto = Producto.objects.get(id=producto_id)
            proveedor = Proveedor.objects.get(id=proveedor_id)

            # 🧮 Cálculos (precio_compra = total de la compra)
            subtotal = precio_compra  # el usuario ya ingresa el total
            precio_unitario = subtotal / cantidad  # cálculo automático
            iva_13 = subtotal * Decimal('0.13')
            total = subtotal + iva_13

            # Generar comprobantes únicos
            comprobante_numero = generar_comprobante_unico()
            registro_nrc = generar_registro_unico()

            # 💾 Guardar la compra
            Compra.objects.create(
                fecha=now(),
                comprobante_numero=comprobante_numero,
                registro_nrc=registro_nrc,
                proveedor=proveedor.nombre,
                precio_unitario=precio_unitario,
                compras_gravadas=subtotal,
                iva_13=iva_13,
                total=total
            )

            # 📦 Registrar movimiento de inventario
            Inventario.objects.create(
                producto=producto,
                tipo='Entrada',
                cantidad=cantidad,
                descripcion=f"Compra registrada de {proveedor.nombre}"
            )

            # 🧰 Actualizar el precio del producto (solo si cambia)
            if producto.precio_unitario != precio_unitario:
                producto.precio_unitario = precio_unitario
                producto.save()

            # 🧾 Crear DTE vinculado
            empresa = Empresa.objects.first()
            cliente = Cliente.objects.first()
            if empresa and cliente:
                DTE.objects.create(
                    empresa=empresa,
                    cliente=cliente,
                    tipo_dte='03',
                    numero_control=comprobante_numero,
                    subtotal=subtotal,
                    iva=iva_13,
                    total=total,
                    codigo_generacion=f"Compra a proveedor: {proveedor.nombre}",
                    estado='Activo',
                    fecha_emision=now()
                )

            messages.success(request, f"✅ Compra registrada correctamente: {cantidad} unidades por ${subtotal} (Precio unitario ${precio_unitario:.2f})")
            return redirect('libro_compras')

        except Exception as e:
            messages.error(request, f"❌ Error al registrar la compra: {str(e)}")
            return redirect('registrar_compra')

    # 🔽 Mostrar formulario
    productos = Producto.objects.all().order_by('descripcion')
    proveedores = Proveedor.objects.all().order_by('nombre')

    return render(request, 'Facturacion/form_compra.html', {
        'productos': productos,
        'proveedores': proveedores
    })

# ======================================================
# 📘 LIBRO DE COMPRAS
# ======================================================

@login_required
def libro_compras(request):
    query = request.GET.get('q', '').strip()

    compras_qs = Compra.objects.all().order_by('-fecha', '-id')

    if query:
        compras_qs = compras_qs.filter(
            Q(comprobante_numero__icontains=query) |
            Q(registro_nrc__icontains=query)
        )

    # Mostrar solo las últimas 10
    compras_qs = compras_qs[:10]

    compras = [
        {
            "fecha": c.fecha.strftime("%d/%m/%Y"),
            "comprobante": c.comprobante_numero,
            "registro": c.registro_nrc,
            "proveedor": c.proveedor,
            "precio_unitario": float(c.precio_unitario),
            "gravadas": float(c.compras_gravadas),
            "iva": float(c.iva_13),
            "total": float(c.total),
        }
        for c in compras_qs
    ]

    return render(request, 'Facturacion/libro_compras.html', {
        'compras': compras,
        'query': query
    })

@login_required
def exportar_libro_compras_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from decimal import Decimal

    # 📘 Obtener los datos igual que en la vista libro_compras
    compras = [
        {
            "fecha": c.fecha.strftime("%d/%m/%Y"),
            "comprobante": c.comprobante_numero,
            "registro": c.registro_nrc,
            "proveedor": c.proveedor,
            "precio_unitario": float(c.precio_unitario) if hasattr(c, "precio_unitario") else 0.0,
            "gravadas": float(c.compras_gravadas),
        }
        for c in Compra.objects.all().order_by('-fecha')
    ]

    # 📗 Crear archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro de Compras"

    # Encabezados
    encabezados = [
        "N°", "Fecha", "Comprobante N°", "Registro N°", "Proveedor",
        "Precio Unitario", "Compras Gravadas", "IVA 13%", "Total"
    ]
    ws.append(encabezados)

    # 🟦 Estilos de encabezado
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # 🧱 Borde de celdas
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 🧮 Agregar filas de datos
    for i, compra in enumerate(compras, start=2):
        fila = [
            i - 1,
            compra["fecha"],
            compra["comprobante"],
            compra["registro"],
            compra["proveedor"],
            compra["precio_unitario"],
            compra["gravadas"],
            f"=G{i}*0.13",   # IVA 13% (columna G)
            f"=G{i}+H{i}"    # Total = Gravadas + IVA
        ]
        ws.append(fila)

    # 🔲 Aplicar bordes y alineación a todas las celdas
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

    # 📏 Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # 💾 Preparar respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Libro_Compras.xlsx"'
    wb.save(response)
    return response



# ======================================================
# 📗 LIBRO DE VENTAS (ajustado a zona horaria local)
# ======================================================

@login_required
def libro_ventas(request):
    """
    Genera el libro de ventas agrupado por día (zona horaria local),
    mostrando los 10 días más recientes.
    """
    from collections import defaultdict
    from django.utils.timezone import localtime
    from datetime import datetime

    dtes = DTE.objects.filter(tipo_dte='01', estado='Activo').order_by('fecha_emision')

    ventas_por_dia = defaultdict(lambda: {'ventas_gravadas': 0, 'total': 0, 'emitido_del': None, 'emitido_al': None})

    for dte in dtes:
        # 🔹 Convierte a hora local antes de agrupar
        fecha_local = localtime(dte.fecha_emision)
        dia = fecha_local.strftime("%d/%m/%Y")

        ventas_por_dia[dia]['ventas_gravadas'] += float(dte.subtotal)
        ventas_por_dia[dia]['total'] += float(dte.total)

        if not ventas_por_dia[dia]['emitido_del']:
            ventas_por_dia[dia]['emitido_del'] = dte.numero_control
        ventas_por_dia[dia]['emitido_al'] = dte.numero_control

    # Convertir a lista y ordenar por fecha descendente
    ventas = [
        {
            'dia': dia,
            'emitido_del': data['emitido_del'],
            'emitido_al': data['emitido_al'],
            'ventas_gravadas': round(data['ventas_gravadas'], 2),
            'total': round(data['total'], 2),
        }
        for dia, data in ventas_por_dia.items()
    ]

    # 🧭 Ordenar de más reciente a más antigua
    ventas.sort(key=lambda x: datetime.strptime(x['dia'], "%d/%m/%Y"), reverse=True)

    # Mostrar solo las últimas 10
    ventas = ventas[:10]

    print(f"📘 DTE encontrados: {len(dtes)}, días únicos: {len(ventas)}")  # para depurar

    return render(request, 'Facturacion/libro_ventas.html', {'ventas': ventas})


# ======================================================
# 🧑‍💼 USUARIOS
# ======================================================

@rol_requerido(['Administrador'])
def lista_usuarios(request):
    usuarios = User.objects.all().select_related('perfil').order_by('-date_joined')[:10]
    return render(request, 'Facturacion/usuarios.html', {'usuarios': usuarios})

@rol_requerido(['Administrador'])
def editar_usuario(request, user_id):
    usuario = get_object_or_404(User, id=user_id)
    if request.method == "POST":
        usuario.username = request.POST.get("username")
        usuario.email = request.POST.get("email")
        usuario.save()
        messages.success(request, "Usuario actualizado correctamente.")
        return redirect('lista_usuarios')
    return render(request, 'Facturacion/editar_usuario.html', {'usuario': usuario})


@rol_requerido(['Administrador'])
def eliminar_usuario(request, user_id):
    usuario = get_object_or_404(User, id=user_id)
    usuario.delete()
    messages.success(request, "Usuario eliminado correctamente.")
    return redirect('lista_usuarios')

def _generar_numero_control(tipo):
    """Genera un número de control ficticio único."""
    return f"DTE-{tipo}-" + uuid.uuid4().hex[:8].upper()


def _generar_qr_datauri(text):
    """Genera un QR embebido en formato base64."""
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_M)
    qr.add_data(text)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    bio = io.BytesIO()
    img.save(bio, format="PNG")
    return f"data:image/png;base64,{base64.b64encode(bio.getvalue()).decode()}"

@csrf_exempt
@login_required
def modal_dte(request, tipo):
    """Crea y carga el DTE con datos del catálogo (producto, cliente, totales)"""
    from .models import Empresa, Cliente, DTE
    from django.utils import timezone

    empresa = Empresa.objects.first()
    if not empresa:
        empresa = Empresa.objects.create(
            nombre="Omnigest S.A. de C.V.",
            nit="0614-240325-001-1",
            nrc="123456-7",
            direccion="San Salvador Centro frente a Salvador del Mundo",
            telefono="2290-9444",
            correo="info@omnigest.com",
            representante_legal="Sucursal Salvador del Mundo"
        )

    cliente = Cliente.objects.first()
    producto = None

    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        venta = data.get('producto', {})
        cliente_id = data.get('clienteId')
        nuevo_cliente = data.get('nuevoCliente')

        # Buscar o crear cliente
        if cliente_id:
            cliente = Cliente.objects.filter(id=cliente_id).first()
        elif nuevo_cliente:
            cliente = Cliente.objects.create(
                nombre=nuevo_cliente,
                correo=data.get('correo', ''),
                direccion=data.get('direccion', '')
            )

        if venta:
            producto = {
                'nombre': venta.get('nombre'),
                'cantidad': venta.get('cantidad'),
                'precio': venta.get('precio'),
                'subtotal': venta.get('subtotal'),
                'iva': venta.get('iva'),
                'total': venta.get('total'),
            }

    dte = DTE.objects.create(
        empresa=empresa,
        cliente=cliente,
        tipo_dte=tipo,
        numero_control=f"DTE-{tipo}-{uuid.uuid4().hex[:8].upper()}",
        fecha_emision=timezone.now(),
        condicion_pago="Contado",
        subtotal=producto['subtotal'] if producto else 0,
        iva=producto['iva'] if producto else 0,
        total=producto['total'] if producto else 0,
        codigo_generacion="(Asignado por MH)",
        sello_recepcion="(Pendiente)",
        estado="Activo"
    )

    plantillas_dte = {
        '01': 'Facturacion/factura01.html',
        '03': 'Facturacion/ccf03.html',
        '05': 'Facturacion/notaCredito05.html',
        '06': 'Facturacion/notaDebito06.html',
        '07': 'Facturacion/retencion07.html',
        '11': 'Facturacion/liquidacion11.html',
    }

    return render(request, plantillas_dte.get(tipo, 'Facturacion/factura01.html'), {
        'tipo_dte': tipo,
        'dte': dte,
        'empresa': empresa,
        'cliente': cliente,
        'producto': producto
    })

@login_required
def ver_dte(request, dte_id):
    """
    Muestra el DTE (desde el botón Editar o QR), con productos incluidos.
    Si el usuario es Administrador, el modal será editable.
    """
    from .models import DTE, DetalleDTE, Perfil

    dte = get_object_or_404(DTE, id=dte_id)
    empresa = dte.empresa
    cliente = dte.cliente

    # 🔹 Obtener los productos asociados
    detalles = DetalleDTE.objects.filter(dte=dte).select_related('producto')

    # 🔹 Verificar si el usuario es Administrador
    perfil = Perfil.objects.filter(user=request.user).first()
    es_admin = perfil and perfil.rol.lower() == 'administrador'

    # 🔹 Generar QR
    qr_text = f"{dte.numero_control}|{empresa.nit}|{cliente.nit}|{dte.total:.2f}"
    qr_data = _generar_qr_datauri(qr_text)

    # 🔹 Plantilla según tipo
    plantillas_dte = {
        '01': 'Facturacion/factura01.html',
        '03': 'Facturacion/ccf03.html',
        '05': 'Facturacion/notaCredito05.html',
        '06': 'Facturacion/notaDebito06.html',
        '07': 'Facturacion/retencion07.html',
        '11': 'Facturacion/liquidacion11.html',
    }
    template = plantillas_dte.get(dte.tipo_dte, 'Facturacion/factura01.html')

    # 🔹 Contexto
    contexto = {
        'dte': dte,
        'empresa': empresa,
        'cliente': cliente,
        'detalles': detalles,
        'qr_data': qr_data,
        'editable': es_admin,  # habilita edición si es admin
    }

    return render(request, template, contexto)


from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.contrib.auth import authenticate
from .models import Perfil
import json

@csrf_exempt
def validar_admin(request):
    """
    Valida las credenciales de un usuario con rol 'Administrador'.
    Compatible con el modelo Perfil.
    """
    try:
        data = json.loads(request.body.decode('utf-8'))
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()

        user = authenticate(username=username, password=password)
        if not user:
            return JsonResponse({'success': False, 'error': 'Usuario o contraseña incorrectos'})

        perfil = Perfil.objects.filter(user=user).first()
        if not perfil:
            return JsonResponse({'success': False, 'error': 'El usuario no tiene perfil asignado'})

        if perfil.rol.lower() == 'administrador':
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': 'No tiene permisos de administrador'})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def editar_dte(request, numero_control):
    """
    Carga el DTE para edición (solo si usuario tiene permisos de administrador).
    """
    from .models import Empresa, Cliente, DTE, DetalleDTE

    dte = get_object_or_404(DTE, numero_control=numero_control)
    detalles = dte.detalles.all()

    if request.method == 'POST':
        # Actualiza los campos
        dte.cliente.nombre = request.POST.get('cliente', dte.cliente.nombre)
        dte.cliente.direccion = request.POST.get('direccion', dte.cliente.direccion)
        dte.cliente.correo = request.POST.get('correo', dte.cliente.correo)
        dte.cliente.save()

        dte.condicion_pago = request.POST.get('condicion_pago', dte.condicion_pago)
        dte.estado = request.POST.get('estado', dte.estado)
        dte.total = request.POST.get('total', dte.total)
        dte.save()

        messages.success(request, "✅ Documento actualizado correctamente.")
        return JsonResponse({'success': True})

    return render(request, 'Facturacion/editar_dte.html', {
        'dte': dte,
        'empresa': dte.empresa,
        'cliente': dte.cliente,
        'detalles': detalles,
    })

@csrf_exempt
@login_required
def actualizar_dte(request, numero_control):
    """
    Actualiza un DTE y recalcula totales para reflejarlo en reportes y libros.
    """
    from .models import DTE, DetalleDTE
    import json

    try:
        dte = DTE.objects.filter(numero_control=numero_control).select_related('cliente', 'empresa').first()
        if not dte:
            return JsonResponse({'success': False, 'error': 'No se encontró el documento.'})

        data = json.loads(request.body.decode('utf-8'))

        # 🔹 Actualizar campos básicos (solo si vienen en el request)
        if 'estado' in data:
            dte.estado = data['estado']
        if 'condicion_pago' in data:
            dte.condicion_pago = data['condicion_pago']
        if 'total' in data:
            try:
                dte.total = Decimal(str(data['total'])).quantize(Decimal('0.01'))
            except Exception:
                dte.total = Decimal('0.00')

        # 🔹 Actualizar cliente si cambió
        if 'cliente' in data:
            dte.cliente.nombre = data['cliente']
        if 'correo' in data:
            dte.cliente.correo = data['correo']
        if 'direccion' in data:
            dte.cliente.direccion = data['direccion']
        dte.cliente.save()

        # 🔹 Recalcular subtotal e IVA
        if hasattr(dte, 'detalles') and dte.detalles.exists():
            subtotal = sum(Decimal(det.total_item) for det in dte.detalles.all())
        else:
            subtotal = (dte.total / Decimal('1.13')).quantize(Decimal('0.01'))

        iva = (subtotal * Decimal('0.13')).quantize(Decimal('0.01'))
        dte.subtotal = subtotal
        dte.iva = iva
        dte.total = subtotal + iva
        dte.fecha_emision = timezone.now()
        dte.save()

        # 🔹 Respuesta de confirmación
        return JsonResponse({
            'success': True,
            'msg': '✅ Documento actualizado y sincronizado correctamente.',
            'total': f"{dte.total:.2f}",
            'iva': f"{dte.iva:.2f}",
            'subtotal': f"{dte.subtotal:.2f}"
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
def buscar_dte(request):
    """Busca documentos DTE por cliente o número de control"""
    q = request.GET.get("q", "").strip()
    resultados = []

    if q:
        # 🔹 Filtramos por cliente o número de control
        dtes = (
            DTE.objects.filter(
                Q(cliente__nombre__icontains=q) | Q(numero_control__icontains=q)
            )
            .select_related("cliente")
            .order_by("-fecha_emision")[:10]  # ✅ Solo los 10 más recientes
        )

        resultados = [
            {
                "id": d.id,
                "tipo_dte": d.get_tipo_dte_display(),
                "numero_control": d.numero_control,
                "cliente": d.cliente.nombre,
                # ✅ Convertir la fecha a zona local (El Salvador) y formatearla
                "fecha": localtime(d.fecha_emision).strftime("%Y-%m-%d %H:%M"),
                "total": str(round(d.total, 2)),
                "estado": d.estado,
            }
            for d in dtes
        ]

    return JsonResponse({"resultados": resultados})

@login_required
def catalogo_productos(request):
    """
    Muestra los productos en stock, permite seleccionar cantidad, cliente y generar un DTE.
    """
    from .models import Producto, Cliente
    productos = Producto.objects.all().order_by('descripcion')
    clientes = Cliente.objects.all().order_by('nombre')
    return render(request, 'Facturacion/catalogo_productos.html', {
        'productos': productos,
        'clientes': clientes
    })

@login_required
def registrar_venta(request):
    if request.method == "POST":
        try:
            producto_id = request.POST.get('producto_id')
            cantidad = request.POST.get('cantidad')
            cliente_id = request.POST.get('cliente_id')
            nuevo_cliente = request.POST.get('nuevo_cliente')
            correo = request.POST.get('correo')
            direccion = request.POST.get('direccion')

            print("🧩 Datos recibidos:", producto_id, cantidad, cliente_id, nuevo_cliente)

            if not producto_id or not cantidad:
                return JsonResponse({'status': 'error', 'mensaje': 'Faltan datos del producto o cantidad.'}, status=400)

            producto = get_object_or_404(Producto, id=int(producto_id))
            cantidad = int(cantidad)

            # 🧾 Buscar o crear cliente
            if cliente_id:
                cliente = get_object_or_404(Cliente, id=int(cliente_id))
            elif nuevo_cliente:
                cliente = Cliente.objects.create(
                    nombre=nuevo_cliente,
                    correo=correo or "sin_correo@omnigest.com",
                    direccion=direccion or "Sin dirección"
                )
                print(f"👤 Nuevo cliente creado: {cliente.nombre}")
            else:
                cliente, _ = Cliente.objects.get_or_create(
                    nombre="Consumidor Final",
                    defaults={"correo": "consumidor@omnigest.com"}
                )

            # ⚙️ Empresa por defecto
            empresa = Empresa.objects.first()
            if not empresa:
                empresa = Empresa.objects.create(
                    nombre="OMNIGEST S.A. de C.V.",
                    nit="0614-000000-001-0",
                    nrc="000000-0",
                    direccion="San Salvador",
                    telefono="2222-1111",
                    correo="info@omnigest.com"
                )

            # ⚠️ Validar stock
            if producto.inventario < cantidad:
                return JsonResponse({'status': 'error', 'mensaje': 'Stock insuficiente'}, status=400)

            # 🧮 Cálculos
            subtotal = producto.precio_unitario * Decimal(cantidad)
            iva = subtotal * Decimal('0.13')
            total = subtotal + iva

            # 📦 Descontar inventario (una sola vez)
            producto.inventario -= cantidad
            producto.save()

            # 🧾 Crear DTE
            dte = DTE.objects.create(
                empresa=empresa,
                cliente=cliente,
                tipo_dte='01',
                numero_control=f"DTE-01-{now().strftime('%Y%m%d%H%M%S')}",
                subtotal=subtotal,
                iva=iva,
                total=total,
                codigo_generacion=f"VENTA-{now().strftime('%H%M%S')}",
                estado='Activo',
                fecha_emision=now()
            )

            # 📦 Registrar movimiento de inventario
            Inventario.objects.create(
                producto=producto,
                tipo='Salida',
                cantidad=cantidad,
                descripcion=f"Venta a {cliente.nombre}"
            )

            return JsonResponse({
                "status": "ok",
                "msg": f"✅ Venta registrada exitosamente para {cliente.nombre}.",
                "nuevo_stock": producto.inventario
            })

        except Exception as e:
            print("❌ Error en registrar_venta:", e)
            return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'mensaje': 'Método no permitido'}, status=405)




